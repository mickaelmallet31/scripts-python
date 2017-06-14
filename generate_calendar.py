#/bin/env python

import calendar
import sys
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Border, Side, PatternFill, Font, GradientFill, Alignment
from datetime import date
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule


#
# Doc: https://openpyxl.readthedocs.io/en/default/
#

# parameters to define the calendar
start_month = 1
end_month = 12
year = 2018

# Create the workbook and grab the active worksheet
wb = Workbook()
ws = wb.active

# Rows can also be appended
line = start_line = 1
current_col = start_col = 1
ws.append([""])
ws.append(["Week day"])
ws.append(["Date"])
line += 3

# To be updated with the moving dates
list_dayoff = [
    '1-1',
    '1-5',
    '8-5',
    # Lundi de Paques
    # Jeudi de l'Ascension
    # Lundi de Pentecote
    '14-7',
    '15-8',
    '1-11',
    '11-11',
    '25-12',
]

list_weekday = [
    'M',
    'T',
    'W',
    'T',
    'F',
    'S',
    'S',
]

# Style definition
style_weekend = Font(color=colors.RED)
alignment_month = Alignment(horizontal="center", vertical="center")
fill_month = PatternFill("solid", fgColor=colors.BLUE)
fill_weekend = PatternFill("solid", fgColor="DDDDDD")
fill_week = PatternFill("solid", fgColor="CCCCCC")
fill_team = PatternFill("solid", fgColor="81DF8F")
fill_dayoff = PatternFill("solid", fgColor="D3191C")
fill_takenday = PatternFill(start_color='DFBC81', end_color='DFBC81', fill_type='solid')
font_month = Font(b=True, color="FFFFFF")

# Freeze the panel
c = ws['C4']
ws.freeze_panes = c

# Display the list of the team members
list_team = []
if len(sys.argv) != 2:
    print '\nERROR:\n\tThanks to provide a text file with the team member (one name per line)'
    exit(1)

for member in open(sys.argv[1], 'r'):
    member = member.strip()
    list_team.append(member)

for member in sorted(list_team):
    ws.cell(row=line, column=current_col, value=member).fill = fill_team
    line += 1
ws.column_dimensions[get_column_letter(current_col)].width = 31.29
current_col += 1
ws.column_dimensions[get_column_letter(current_col)].width = 3

end_line = line

# For each month
for month in range(start_month, end_month+1):
    # Go to the next column
    current_col += 1

    # Get the number of days
    nb_days = calendar.monthrange(year, month)[1]

    # Display the name of the month
    ws.cell(row=start_line, column=current_col, value=calendar.month_name[month])
    ws.merge_cells(start_row=start_line, start_column=current_col, end_row=start_line, end_column=current_col+nb_days)
    ws.cell(row=start_line, column=current_col).alignment = alignment_month
    ws.cell(row=start_line, column=current_col).fill = fill_month
    ws.cell(row=start_line, column=current_col).font = font_month

    # Display each day number of the current month
    for num_day in range(1, nb_days+1):
        num_weekday = date(year, month, num_day).weekday()
        curr_col = get_column_letter(num_day+current_col-1)
        ws.cell(row=start_line+1, column=num_day+current_col-1, value=list_weekday[num_weekday])
        ws.cell(row=start_line+2, column=num_day+current_col-1, value=num_day)
        ws.column_dimensions[curr_col].width = 3
        ws.cell(row=start_line+1, column=num_day+current_col-1).alignment = alignment_month
        ws.cell(row=start_line+2, column=num_day+current_col-1).alignment = alignment_month
        # Display weekend days in different color
        if num_weekday > 4:
            for row in range(start_line+1, start_line+3+len(list_team)):
                ws.cell(row=row, column=num_day+current_col-1).fill = fill_weekend
        else:
            # if the day is a dayoff then display in another color
            date_month = '{}-{}'.format(num_day, month)
            if date_month in list_dayoff:
                for row in range(start_line+1, start_line+3+len(list_team)):
                    ws.cell(row=row, column=num_day+current_col-1).fill = fill_dayoff
            else:
                ws.cell(row=start_line+1, column=num_day+current_col-1).fill = fill_week
                ws.cell(row=start_line+2, column=num_day+current_col-1).fill = fill_week

    # Display the total column
    current_col += num_day
    ws.merge_cells(start_row=start_line+1, start_column=current_col, end_row=start_line+2, end_column=current_col)
    ws.cell(row=start_line+1, column=current_col, value='Total')
    for row in range(start_line+3, start_line+3+len(list_team)):
        ws.cell(row=row, column=current_col, value='=SUM({}{}:{}{})'.format(get_column_letter(current_col-num_day), row, get_column_letter(current_col-1), row))
    ws.column_dimensions[get_column_letter(current_col)].width = 5

    # Blank column between 2 months
    current_col += 1
    ws.column_dimensions[get_column_letter(current_col)].width = 3

# Format if cell is between 'formula'
ws.conditional_formatting.add('{}{}:{}{}'.format(get_column_letter(3), start_line+3, get_column_letter(current_col), start_line+3+len(list_team)), CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=fill_takenday))

# Save the file
wb.save("calendar.xlsx")
